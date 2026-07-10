import uproot
from openpyxl import Workbook
from openpyxl.styles import Font

def list_histograms(root_file_path):
    """Returns { 'hist_name': entries, ... } for a file."""
    histograms = {}
    try:
        with uproot.open(root_file_path) as root_file:
            for hist_name, hist in root_file.items():
                if isinstance(hist, uproot.behaviors.TH1.Histogram):
                    # Process the histogram name
                    if ';' in hist_name:
                        hist_name = hist_name.split(';')[0]  # Remove trailing ;1 if present
                    
                    # Parse the original name
                    parts = hist_name.split('_')
                    if len(parts) >= 4 and 'bin' in parts[0] and ('Pass' in hist_name or 'Fail' in hist_name):
                        # Extract components
                        bin_num = parts[0].replace('bin', '')
                        pt_range = parts[2].replace('p', '.').replace('To', ' - ')
                        pass_fail = 'pass' if 'Pass' in hist_name else 'fail'
                        
                        # Create new name format
                        new_name = f"bin_{bin_num}: {pt_range} ({pass_fail})"
                        histograms[new_name] = hist.values().sum()
                    else:
                        # Keep original name if format doesn't match
                        histograms[hist_name] = hist.values().sum()
    except Exception as e:
        print(f"Error processing {root_file_path}: {e}")
    return histograms

def save_to_excel(file_mapping, output_excel="old_blp_histogram_entries.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Histogram Entries"

    # Headers
    headers = ["Histogram Name"] + list(file_mapping.keys())
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header).font = Font(bold=True)

    # Load all data
    data = {}
    all_hist_names = set()
    for col_name, file_path in file_mapping.items():
        data[col_name] = list_histograms(file_path)
        all_hist_names.update(data[col_name].keys())

    # Sort histogram names alphabetically (this will group pass/fail pairs)
    sorted_hist_names = sorted(all_hist_names)

    # Write data
    for row_num, hist_name in enumerate(sorted_hist_names, 2):
        ws.cell(row=row_num, column=1, value=hist_name)
        for col_num, col_name in enumerate(file_mapping.keys(), 2):
            if hist_name in data[col_name]:
                ws.cell(row=row_num, column=col_num, value=data[col_name][hist_name])

    # Auto-adjust columns
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    wb.save(output_excel)
    print(f"Data saved to {output_excel}")

if __name__ == "__main__":
    file_mapping = {
                        
        "DATA_Barrel_1": "/uscms/home/hortua/nobackup/egamma-tnp/examples/nanoaod_filters_custom/blp2/DATA_2023D/get_1d_pt_eta_phi_tnp_histograms_1/DATA_2_23D_histos_pt_barrel_1.root",
        "DATA_Barrel_2": "/uscms/home/hortua/nobackup/egamma-tnp/examples/nanoaod_filters_custom/blp2/DATA_2023D/get_1d_pt_eta_phi_tnp_histograms_1/DATA_2_23D_histos_pt_barrel_2.root",
        "DATA_Endcap": "/uscms/home/hortua/nobackup/egamma-tnp/examples/nanoaod_filters_custom/blp2/DATA_2023D/get_1d_pt_eta_phi_tnp_histograms_1/DATA_2_23D_histos_pt_endcap.root",
        "MC_Barrel_1": "/uscms/home/hortua/nobackup/egamma-tnp/examples/nanoaod_filters_custom/blp2/MC_DY_2023/get_1d_pt_eta_phi_tnp_histograms_1/MC_2_23D_histos_pt_barrel_1.root",
        "MC_Barrel_2": "/uscms/home/hortua/nobackup/egamma-tnp/examples/nanoaod_filters_custom/blp2/MC_DY_2023/get_1d_pt_eta_phi_tnp_histograms_1/MC_2_23D_histos_pt_barrel_2.root",
        "MC_Endcap": "/uscms/home/hortua/nobackup/egamma-tnp/examples/nanoaod_filters_custom/blp2/MC_DY_2023/get_1d_pt_eta_phi_tnp_histograms_1/MC_2_23D_histos_pt_endcap.root"
    }
    save_to_excel(file_mapping)

      #  "DATA_Barrel_1": "/uscms/home/hortua/nobackup/egamma-tnp/examples/nanoaod_filters_custom/new_blp/DATA_2023D/get_1d_pt_eta_phi_tnp_histograms_1/DATA_23D_histos_pt_barrel_1.root",
      #  "DATA_Barrel_2": "/uscms/home/hortua/nobackup/egamma-tnp/examples/nanoaod_filters_custom/new_blp/DATA_2023D/get_1d_pt_eta_phi_tnp_histograms_1/DATA_23D_histos_pt_barrel_2.root",
      #  "DATA_Endcap": "/uscms/home/hortua/nobackup/egamma-tnp/examples/nanoaod_filters_custom/new_blp/DATA_2023D/get_1d_pt_eta_phi_tnp_histograms_1/DATA_23D_histos_pt_endcap.root",
      #  "MC_Barrel_1": "/uscms/home/hortua/nobackup/egamma-tnp/examples/nanoaod_filters_custom/new_blp/MC_DY_2023/get_1d_pt_eta_phi_tnp_histograms_1/MC_DY_23D_histos_pt_barrel_1.root",
      #  "MC_Barrel_2": "/uscms/home/hortua/nobackup/egamma-tnp/examples/nanoaod_filters_custom/new_blp/MC_DY_2023/get_1d_pt_eta_phi_tnp_histograms_1/MC_DY_23D_histos_pt_barrel_2.root",
      #  "MC_Endcap": "/uscms/home/hortua/nobackup/egamma-tnp/examples/nanoaod_filters_custom/new_blp/MC_DY_2023/get_1d_pt_eta_phi_tnp_histograms_1/MC_DY_23D_histos_pt_endcap.root"